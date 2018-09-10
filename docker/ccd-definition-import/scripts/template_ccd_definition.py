#!/usr/bin/env python3
from openpyxl import load_workbook
import os
import sys

if len(sys.argv) != 3:
  print('usage: ./template_ccd_definition.py <template-file-location> <output-file-location>', file=sys.stderr)
  sys.exit(1)

template_file_location = sys.argv[1]
output_file_location = sys.argv[2]

bulk_scan_orchestrator_base_url = os.environ['BULK_SCAN_ORCHESTRATOR_BASE_URL']

wb = load_workbook(template_file_location)
bulk_scan_template = '${BULK_SCAN_ORCHESTRATOR_BASE_URL}'

def template_sheet(ws, callback_columns):
  print('Updating CCD sheet ' + ws.title)
  count_of_cells_updated = 0
  for column in callback_columns:
    for cell in ws[column]:
      if cell.value is not None and bulk_scan_template in cell.value:
        cell.value = cell.value.replace(bulk_scan_template, bulk_scan_orchestrator_base_url)
        count_of_cells_updated += 1
  print('Successfully updated BSO %s, %d cells were updated' % (ws.title, count_of_cells_updated))

case_event_ws = wb['CaseEvent']
case_event_callback_columns = ['J', 'L', 'N']
template_sheet(case_event_ws, case_event_callback_columns)

wb.save(output_file_location)
